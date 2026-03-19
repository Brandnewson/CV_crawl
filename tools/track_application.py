"""Append one row to the Excel application tracker.

Usage:
    uv run --project "C:/Code/CV_crawl" \\
        python tools/track_application.py \\
        --job-id 42 \\
        --company "Reflection AI" \\
        --job-title "Forward Deployed Engineer" \\
        --city "New York, US" \\
        --description "..." \\
        --job-url "https://..."
"""

from __future__ import annotations

import argparse
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

TRACKER_PATH = Path(
    r"C:\Users\brans\OneDrive - University of Leeds\GraduateJobHunting\applications_tracker.xlsx"
)
COLUMNS = ["Date Applied", "Company", "Job Title", "City", "Job Description", "Job URL"]


def _auto_size_columns(ws) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(cell.value or "")) for cell in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def track(
    job_id: int,
    company: str,
    job_title: str,
    city: str,
    description: str,
    job_url: str,
) -> None:
    if TRACKER_PATH.exists():
        wb = openpyxl.load_workbook(TRACKER_PATH)
        ws = wb.active
    else:
        TRACKER_PATH.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(COLUMNS)

    ws.append([
        date.today().isoformat(),
        company,
        job_title,
        city,
        description[:2000],
        job_url,
    ])
    _auto_size_columns(ws)
    wb.save(TRACKER_PATH)
    print(f"[tracker] Appended row for {company} - {job_title} -> {TRACKER_PATH}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Track job application in Excel")
    parser.add_argument("--job-id", type=int, required=True)
    parser.add_argument("--company", required=True)
    parser.add_argument("--job-title", required=True)
    parser.add_argument("--city", default="")
    parser.add_argument("--description", default="")
    parser.add_argument("--job-url", default="")
    args = parser.parse_args()

    track(
        job_id=args.job_id,
        company=args.company,
        job_title=args.job_title,
        city=args.city,
        description=args.description,
        job_url=args.job_url,
    )


if __name__ == "__main__":
    main()
